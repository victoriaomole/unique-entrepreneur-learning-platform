import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Base path
BASE_DIR = r"C:\Users\victo\kickoff_project"
TARGET_DIR = os.path.join(BASE_DIR, "01_Phase1_Discovery_and_Scoping")
os.makedirs(TARGET_DIR, exist_ok=True)

FILE_PATH = os.path.join(TARGET_DIR, "Discovery_Interview_Log.xlsx")

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Discovery Interviews"

# Columns
columns = [
    "Interview ID",
    "Stakeholder Type",
    "Organisation / School",
    "Interviewee Name",
    "Country",
    "Sector",
    "Date",
    "Key Themes",
    "Pain Points",
    "Opportunities",
    "Follow-ups",
    "Satisfaction (1-5)",
    "Priority Level",
    "Linked File"
]

# Header style
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E375A", end_color="1E375A", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write header row
for col_idx, col_name in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align

# Sample realistic rows
rows = [
    [
        "INT-001",
        "Org Admin",
        "GreenCycle Cooperative",
        "Mrs. Olufunke Adewale",
        "Nigeria",
        "Recycling Cooperative",
        "08-Nov-2025",
        "Membership enrolment, learning access, reporting",
        "Manual paper records; no centralised tracking; difficulty monitoring training completion.",
        "Introduce member portal with online training, attendance, and simple reports.",
        "Schedule demo of proposed platform for cooperative leadership.",
        3,
        "High",
        r"Stakeholder_Interview_Notes\OrgAdmin_Interview_01.docx",
    ],
    [
        "INT-002",
        "Instructor",
        "Manchester AgriHub",
        "Sarah Jones",
        "UK",
        "Agriculture Training",
        "09-Nov-2025",
        "Course creation, media upload, rural learner access",
        "Slow video uploads; learners with poor bandwidth struggle to stream content.",
        "Add adaptive video streaming, downloadable resources, and compressed formats.",
        "Test pilot course with low-bandwidth configuration.",
        4,
        "High",
        r"Stakeholder_Interview_Notes\Instructor_Interview_02.docx",
    ],
    [
        "INT-003",
        "Student",
        "Unique Youth Enterprise Centre",
        "Adeyemi Tunde",
        "Nigeria",
        "Entrepreneurship",
        "10-Nov-2025",
        "Mobile learning, certificates, affordability",
        "Can’t afford data-heavy platforms; progress tracking unclear.",
        "Lightweight mobile-friendly learner interface, offline notes, clear certificate display.",
        "Collect feedback from 20+ youth learners after pilot.",
        5,
        "Medium",
        r"Stakeholder_Interview_Notes\Student_Interview_03.docx",
    ],
    [
        "INT-004",
        "DMP / Compliance",
        "ReGreen Agro Solutions Ltd",
        "Dr. Fiona Clark",
        "UK",
        "Agri Data & Training",
        "10-Nov-2025",
        "GDPR, audit trails, data sharing",
        "No unified view of where learner data sits; weak audit logging.",
        "Implement clear data ownership model, RLS by org, immutable audit logs.",
        "Draft Data Protection Impact Assessment (DPIA) based on proposed architecture.",
        4,
        "High",
        r"Stakeholder_Interview_Notes\DMP_Compliance_Interview_04.docx",
    ],
    [
        "INT-005",
        "Executive Sponsor",
        "Unique Cooperative Society",
        "Mr. Chukwuma Eze",
        "Nigeria",
        "Cooperative & Microfinance",
        "11-Nov-2025",
        "Impact tracking, KPIs, financial sustainability",
        "No structured way to see how training affects loan performance or member growth.",
        "Dashboards linking course completion to member retention and repayment behaviour.",
        "Define KPI framework and reporting cadence for ExCo.",
        3,
        "High",
        r"Stakeholder_Interview_Notes\ExecSponsor_Interview_05.docx",
    ],
    [
        "INT-006",
        "Technical Lead",
        "EcoFarm Learning UK",
        "James Bennett",
        "UK",
        "AgriTech Education",
        "12-Nov-2025",
        "Integration, uptime, connectivity constraints",
        "Rural sites with unstable internet; need secure SSO and simple integrations.",
        "Hybrid-friendly cloud setup, robust API, offline-tolerant client patterns.",
        "Run technical feasibility workshop with platform engineering team.",
        4,
        "Medium",
        r"Stakeholder_Interview_Notes\TechLead_Interview_06.docx",
    ],
]

# Write data rows
for row_idx, row_data in enumerate(rows, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if col_idx in (12,):  # Satisfaction numeric
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in (1, 2, 5, 6, 13):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

# Auto column width (simple)
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

wb.save(FILE_PATH)

print(f"✅ Discovery Interview Log created at:\n{FILE_PATH}")

